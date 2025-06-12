import tkinter as tk
from tkinter import ttk, messagebox
from tkcalendar import DateEntry
from openpyxl import load_workbook, Workbook
from fuzzywuzzy import fuzz
import os
from datetime import datetime

# === Excel 檔案名稱 ===
FOOD_FILE = "foodssugar.xlsx"
RECORD_FILE = "Ruby_records.xlsx"

# === 初始化 ===
def init_excel():
    if not os.path.exists(FOOD_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "食物資料"
        ws.append(["食物名稱", "單位", "碳水化合物", "備註"])
        wb.save(FOOD_FILE)

# === 檢查是否有相似品項 ===
def find_similar_foods(food_name, threshold=80):
    wb = load_workbook(FOOD_FILE)
    ws = wb.active
    matches = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if fuzz.partial_ratio(food_name, row[0]) >= threshold:
            matches.append(row)
    return matches

# === 新增食物資料 ===
def check_and_save_food():
    name = entry_name.get().strip()
    unit = combo_unit.get().strip()
    carb = entry_carb.get().strip().replace(",", ".")  # 新增 replace
    note = entry_note.get().strip()

    if not name or not unit or not carb:
        messagebox.showwarning("錯誤", "請填寫完整資訊")
        return
    try:
        float(carb)
    except:
        messagebox.showwarning("錯誤", "碳水化合物請填數字")
        return

    matches = find_similar_foods(name)
    wb = load_workbook(FOOD_FILE)
    ws = wb.active

    if matches:
        msg = "發現相似食物：\n"
        for row in matches:
            msg += f"- {row[0]} ({row[1]}, {row[2]}g)\n"
        msg += "\n是否要覆蓋第一筆？"
        if messagebox.askyesno("相似品項", msg):
            for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
                if fuzz.partial_ratio(name, row[0].value) >= 80:
                    ws.cell(i, 1, name)
                    ws.cell(i, 2, unit)
                    ws.cell(i, 3, float(carb))
                    ws.cell(i, 4, note)
                    wb.save(FOOD_FILE)
                    messagebox.showinfo("成功", "已覆蓋相似食物")
                    return
    ws.append([name, unit, float(carb), note])
    wb.save(FOOD_FILE)
    messagebox.showinfo("成功", "已新增食物")
    entry_name.delete(0, tk.END)
    combo_unit.set("")
    entry_carb.delete(0, tk.END)
    entry_note.delete(0, tk.END)

# === 刪除食物資料 ===
def delete_food():
    sel = search_listbox.curselection()
    if not sel:
        messagebox.showwarning("請選擇", "請先選擇要刪除的食物")
        return
    selected_name = current_matches[sel[0]][0]
    
    wb = load_workbook(FOOD_FILE)
    ws = wb.active
    deleted = False
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if row[0].value == selected_name:
            ws.delete_rows(i)
            deleted = True
            break
    if deleted:
        wb.save(FOOD_FILE)
        messagebox.showinfo("已刪除", f"已刪除 {selected_name}")
        search_food()  # 重新查詢
    else:
        messagebox.showerror("錯誤", "找不到對應食物")

# === 查詢功能 ===
def search_food():
    search_listbox.delete(0, tk.END)
    global current_matches
    keyword = entry_search.get().strip()
    current_matches = find_similar_foods(keyword)
    if not current_matches:
        messagebox.showinfo("查無資料", "請先新增食物")
        return
    for row in current_matches:
        search_listbox.insert(tk.END, f"{row[0]}｜每{row[1]} 含 {row[2]}g｜{row[3]}")

# === 碳水計算的查詢食物 ===        
def search_food_calc():
    global current_matches
    search_listbox_calc.delete(0, tk.END)
    global current_matches
    keyword = entry_search_calc.get().strip()
    current_matches = find_similar_foods(keyword)
    if not current_matches:
        messagebox.showinfo("查無資料", "請先新增食物")
        return
    for row in current_matches:
        search_listbox_calc.insert(tk.END, f"{row[0]}｜每{row[1]} 含 {row[2]}g｜{row[3]}")

# === 加入計算 ===
def add_to_calc():
    sel = search_listbox.curselection()
    if not sel:
        messagebox.showwarning("請選擇", "請選擇一筆食物")
        return
    try:
        amount = float(entry_amount.get())
        if amount <= 0:
            raise ValueError
    except:
        messagebox.showwarning("錯誤", "攝取量請輸入正數")
        return

    row = current_matches[sel[0]]
    carb = round(float(row[2]) * amount, 2)
    calc_listbox.insert(tk.END, f"{row[0]}｜{amount}{row[1]}｜碳水:{carb}g")
    calc_results.append({"name": row[0], "amount": amount, "unit": row[1], "carb": carb})
    update_total()
    entry_amount.delete(0, tk.END)

def update_total():
    total = round(sum([r["carb"] for r in calc_results]), 2)
    lbl_total.config(text=f"總碳水量：{total:.2f}g")
    
def on_tab_recommend_selected(event):
    # try:
    #     entry_ci.delete(0, tk.END)
    #     entry_ci.insert(0, "從資料表載入的 C/I")

    #     entry_isf.delete(0, tk.END)
    #     entry_isf.insert(0, "從資料表載入的 ISF")
    #     # 其他可補充項目
    # except:
        pass


# === 儲存記錄 ===
def parse_float_input(raw_value, placeholder_keywords):
    raw = raw_value.strip()
    if not raw or any(keyword in raw for keyword in placeholder_keywords):
        return None  # 改成 None 而不是 ""
    try:
        return float(raw)
    except ValueError:
        return None

def save_records():
    date = entry_date.get().strip()
    meal = combo_meal.get().strip()
    if not date or not meal:
        messagebox.showwarning("請填寫", "請輸入日期與餐別")
        return

    try:
        current_glucose = int(entry_current_glucose.get().strip())
    except ValueError:
        messagebox.showwarning("錯誤", "請輸入有效的整數『目前血糖值』")
        return

    target_glucose_raw = entry_target_glucose.get().strip()
    if target_glucose_raw:
        try:
            target_glucose = int(target_glucose_raw)
        except ValueError:
            messagebox.showwarning("錯誤", "『期望血糖值』請輸入整數或留白")
            return
    else:
        target_glucose = None

    total_carb = round(sum([r["carb"] for r in calc_results]), 2) if calc_results else 0.0

    if not calc_results:
        messagebox.showinfo("提醒", "您尚未加入任何食物項目，將僅儲存血糖資訊。")

    def parse_float_input_safe(raw_value, placeholder_keywords):
        raw = raw_value.strip()
        if not raw or any(keyword in raw for keyword in placeholder_keywords):
            return None
        try:
            return float(raw)
        except ValueError:
            return None

    ci_value = parse_float_input_safe(entry_ci.get(), ["從資料表載入", "預設"])
    isf_value = parse_float_input_safe(entry_isf.get(), ["從資料表載入", "預設"])
    c_raise_value = parse_float_input_safe(entry_c_raise.get(), ["從資料表載入", "預設"])

    if ci_value is None or isf_value is None:
        messagebox.showwarning("錯誤", "請填寫 C/I 與 ISF 值")
        return

    if 'last_total_insulin' not in globals() or last_total_insulin is None:
        messagebox.showwarning("錯誤", "請先按下『計算胰島素劑量』後再儲存")
        return

    insulin_carb = last_insulin_carb
    insulin_correction = last_insulin_correction
    total_insulin = last_total_insulin
    recommended_ci = recommended_ci_value if 'recommended_ci_value' in globals() else None

    try:
        if os.path.exists(RECORD_FILE):
            wb = load_workbook(RECORD_FILE)
        else:
            wb = Workbook()

        if "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 1:
            wb.remove(wb["Sheet"])

        # 食物記錄
        if "食物記錄" not in wb.sheetnames:
            ws_food = wb.create_sheet("食物記錄")
            ws_food.append(["日期", "餐別", "食物名稱", "攝取量", "單位", "碳水化合物"])
        else:
            ws_food = wb["食物記錄"]

        if calc_results:
            for item in calc_results:
                ws_food.append([date, meal, item["name"], item["amount"], item["unit"], item["carb"]])
            ws_food.append(["", "", "", "", "總碳水", total_carb])

        # 血糖與胰島素紀錄表
        if "血糖與胰島素紀錄表" not in wb.sheetnames:
            ws_insulin = wb.create_sheet("血糖與胰島素紀錄表")
            ws_insulin.append([
                "日期", "餐別", "總碳水量", "目前血糖值", "期望血糖值",
                "C/I值", "ISF值", "1C升高血糖", "碳水劑量", "矯正劑量", "總胰島素劑量",
                "餐後血糖值", "建議C/I值"
            ])
        else:
            ws_insulin = wb["血糖與胰島素紀錄表"]

        updated = False
        for row in range(2, ws_insulin.max_row + 1):
            existing_date_raw = ws_insulin.cell(row=row, column=1).value
            existing_date = existing_date_raw.strftime("%Y-%m-%d") if isinstance(existing_date_raw, datetime) else str(existing_date_raw).strip()
            existing_meal = str(ws_insulin.cell(row=row, column=2).value).strip()
            if existing_date == date and existing_meal == meal:
                ws_insulin.cell(row=row, column=3).value = total_carb
                ws_insulin.cell(row=row, column=4).value = current_glucose
                ws_insulin.cell(row=row, column=5).value = target_glucose
                ws_insulin.cell(row=row, column=6).value = ci_value
                ws_insulin.cell(row=row, column=7).value = isf_value
                ws_insulin.cell(row=row, column=8).value = c_raise_value
                ws_insulin.cell(row=row, column=9).value = insulin_carb
                ws_insulin.cell(row=row, column=10).value = insulin_correction
                ws_insulin.cell(row=row, column=11).value = total_insulin
                ws_insulin.cell(row=row, column=13).value = recommended_ci
                updated = True
                break

        if not updated:
            ws_insulin.append([
                date, meal, total_carb, current_glucose, target_glucose,
                ci_value, isf_value, c_raise_value,
                insulin_carb, insulin_correction, total_insulin,
                None,  # 餐後血糖值（尚未輸入）
                recommended_ci
            ])

        wb.save(RECORD_FILE)
        messagebox.showinfo("已儲存", f"資料已存入 {RECORD_FILE}")

        # 儲存後詢問是否清除碳水計算
        if messagebox.askyesno("清除確認", "是否要清除目前碳水計算資料？"):
            clear_all()

    except PermissionError:
        messagebox.showerror("儲存失敗", "無法寫入 Excel，請確認是否關閉檔案後再試一次。")


# === 碳水計算的加入計算 ===
def add_to_calc_calc_tab():
    global current_matches
    sel = search_listbox_calc.curselection()
    if not sel:
        messagebox.showwarning("請選擇", "請先選擇一筆食物")
        return

    try:
        amount = float(entry_amount.get())
        if amount <= 0:
            raise ValueError
    except:
        messagebox.showwarning("錯誤", "攝取量請輸入正數")
        return

    row = current_matches[sel[0]]
    carb = round(float(row[2]) * amount, 2)
    calc_listbox.insert(tk.END, f"{row[0]}｜{amount}{row[1]}｜碳水:{carb}g")
    calc_results.append({"name": row[0], "amount": amount, "unit": row[1], "carb": carb})
    update_total()
    entry_amount.delete(0, tk.END)


# === 刪除計算碳水的食物 ===
def delete_selected_calc_item():
    sel = calc_listbox.curselection()
    if not sel:
        messagebox.showwarning("請選擇", "請先選擇要刪除的項目")
        return

    # 移除選擇的項目（從後面開始避免 index 錯位）
    for index in reversed(sel):
        del calc_results[index]
        calc_listbox.delete(index)

    update_total()
    

# === 計算胰島素劑量 ===
def calculate_insulin_dose():
    try:
        total_carb = round(sum([r["carb"] for r in calc_results]), 2)
        ci = float(entry_ci.get().strip())
        isf = float(entry_isf.get().strip())
        current_glucose = int(entry_current_glucose.get().strip())
        target_glucose_raw = entry_target_glucose.get().strip()
        target_glucose = int(target_glucose_raw) if target_glucose_raw else 100  # 預設目標血糖

        if ci <= 0 or isf <= 0:
            raise ValueError

        # 計算劑量
        insulin_carb = total_carb / ci
        insulin_correction = (current_glucose - target_glucose) / isf  # <=== 修正這裡

        # 應用進位規則
        insulin_carb = round_insulin(insulin_carb)
        insulin_correction = round_insulin(insulin_correction)
        total_insulin = round_insulin(insulin_carb + insulin_correction)

        # 顯示結果
        label_insulin_result.config(text=(
            f"碳水劑量: {insulin_carb}U\n"
            f"矯正劑量: {insulin_correction}U\n"
            f"總胰島素劑量: {total_insulin}U"
        ))

        # 儲存到全域變數供儲存用
        global last_insulin_carb, last_insulin_correction, last_total_insulin
        last_insulin_carb = insulin_carb
        last_insulin_correction = insulin_correction
        last_total_insulin = total_insulin

    except:
        messagebox.showwarning("錯誤", "請確認所有數值均為正確格式（C/I、ISF、血糖）")

# === 胰島素劑量進位 ===
def round_insulin(value):
    decimal = value - int(value)
    if decimal <= 0.25:
        return round(int(value) + 0.0, 1)
    elif decimal <= 0.75:
        return round(int(value) + 0.5, 1)
    else:
        return round(int(value) + 1.0, 1)

# === 儲存餐後血糖 ===
def save_post_glucose_only():
    date = entry_date.get().strip()
    meal = combo_meal.get().strip()
    post_glucose_str = entry_post_glucose.get().strip()

    if not date or not meal or not post_glucose_str:
        messagebox.showwarning("錯誤", "請輸入完整的日期、餐別與餐後血糖值")
        return

    try:
        post_glucose = int(post_glucose_str)
    except ValueError:
        messagebox.showwarning("錯誤", "餐後血糖值請輸入整數")
        return

    if os.path.exists(RECORD_FILE):
        wb = load_workbook(RECORD_FILE)
    else:
        wb = Workbook()
    if "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 1:
        wb.remove(wb["Sheet"])

    if "血糖與胰島素紀錄表" not in wb.sheetnames:
        ws = wb.create_sheet("血糖與胰島素紀錄表")
        ws.append([
            "日期", "餐別", "總碳水量", "目前血糖值", "期望血糖值",
            "C/I值", "ISF值", "1C升高血糖", "碳水劑量", "矯正劑量", "總胰島素劑量",
            "餐後血糖值", "建議C/I值"
        ])
    else:
        ws = wb["血糖與胰島素紀錄表"]

    updated = False
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == date and ws.cell(row=row, column=2).value == meal:
            ws.cell(row=row, column=12).value = post_glucose
            updated = True
            break

    if not updated:
        # 其他欄位空白，只填入餐後血糖值
        ws.append([date, meal, "", "", "", "", "", "", "", "", "", post_glucose, ""])

    wb.save(RECORD_FILE)
    messagebox.showinfo("已儲存", f"{date} {meal} 的餐後血糖值已紀錄")
    
# === 計算建議C/I值 ===
def calculate_recommended_ci():
    global recommended_ci_value

    date = entry_date.get().strip()
    meal = combo_meal.get().strip()

    if not date or not meal:
        messagebox.showwarning("錯誤", "請先輸入日期與餐別")
        return

    try:
        post_glucose = int(entry_post_glucose.get().strip())
    except ValueError:
        messagebox.showwarning("錯誤", "請正確輸入餐後血糖值（整數）")
        return

    if not os.path.exists(RECORD_FILE):
        messagebox.showwarning("錯誤", "找不到紀錄檔案")
        return

    wb = load_workbook(RECORD_FILE)
    if "血糖與胰島素紀錄表" not in wb.sheetnames:
        messagebox.showwarning("錯誤", "Excel 中找不到血糖與胰島素紀錄表")
        return

    ws = wb["血糖與胰島素紀錄表"]
    matched = False

    for row in range(2, ws.max_row + 1):
        row_date = str(ws.cell(row=row, column=1).value).strip()
        row_meal = str(ws.cell(row=row, column=2).value).strip()
        if row_date == date and row_meal == meal:
            try:
                total_carb = float(ws.cell(row=row, column=3).value)
                current_glucose = int(ws.cell(row=row, column=4).value)
                isf = float(ws.cell(row=row, column=7).value)
                total_insulin = float(ws.cell(row=row, column=11).value)
            except:
                messagebox.showwarning("錯誤", f"{date} {meal} 的紀錄資訊不完整，請確認已儲存對應資料")
                return

            correction_part = (current_glucose - post_glucose) / isf
            denominator = total_insulin - correction_part

            if denominator <= 0:
                label_ci_recommend.config(text="⚠️ 無法回推有效 C/I（分母為0或負值）")
                recommended_ci_value = ""
                return

            recommended_ci = round(total_carb / denominator, 2)

            # 顯示在畫面上
            label_ci_recommend.config(text=f"🔁 建議調整 C/I 為：{recommended_ci}")
            recommended_ci_value = recommended_ci

            # ✅ 同步寫入 Excel（第13欄）
            ws.cell(row=row, column=13).value = recommended_ci
            wb.save(RECORD_FILE)

            matched = True
            break

    if not matched:
        messagebox.showwarning("錯誤", f"{date} {meal} 的紀錄不存在，請先儲存資料")



# === 帶入前筆C/I值 ===
def load_last_recommended_ci(event=None):
    meal = combo_meal.get().strip()
    date_str = entry_date.get().strip()

    if not meal or not date_str:
        return

    try:
        current_date = datetime.strptime(date_str, "%Y-%m-%d")
    except ValueError:
        messagebox.showwarning("錯誤", "日期格式錯誤")
        return

    if not os.path.exists(RECORD_FILE):
        messagebox.showwarning("錯誤", "找不到記錄檔案")
        return

    wb = load_workbook(RECORD_FILE, data_only=True)
    if "血糖與胰島素紀錄表" not in wb.sheetnames:
        messagebox.showwarning("錯誤", "找不到血糖與胰島素紀錄表")
        return

    ws = wb["血糖與胰島素紀錄表"]
    latest_date = None
    latest_ci = None

    for row in range(2, ws.max_row + 1):
        row_date = ws.cell(row=row, column=1).value
        row_meal = ws.cell(row=row, column=2).value
        row_ci = ws.cell(row=row, column=13).value

        try:
            row_date_obj = datetime.strptime(str(row_date), "%Y-%m-%d")
        except:
            continue

        if row_meal == meal and row_ci and row_date_obj <= current_date:
            if latest_date is None or row_date_obj > latest_date:
                latest_date = row_date_obj
                latest_ci = row_ci

    if latest_ci:
        entry_ci.delete(0, tk.END)
        entry_ci.insert(0, str(round(latest_ci, 2)))
        label_ci_load_info.config(
            text=f"🔁 已自 {latest_date.strftime('%Y-%m-%d')} 的 {meal} 載入建議 C/I 值 {round(latest_ci, 2)}"
        )
    else:
        label_ci_load_info.config(text="❗ 查無建議 C/I 記錄")  # 若查不到則清空提示
        
def clear_insulin_tab():
    # 清空所有輸入欄位
    entry_current_glucose.delete(0, tk.END)
    entry_target_glucose.delete(0, tk.END)
    entry_post_glucose.delete(0, tk.END)
    entry_ci.delete(0, tk.END)
    entry_isf.delete(0, tk.END)
    entry_c_raise.delete(0, tk.END)

    # 日期與餐別可選擇是否清空，這裡也一起清
    entry_date.set_date(datetime.today())
    combo_meal.set("")

    # 清除結果標籤
    label_insulin_result.config(text="")
    label_ci_recommend.config(text="")
    label_ci_load_info.config(text="")

    # 清空全域變數（如需要）
    global last_insulin_carb, last_insulin_correction, last_total_insulin, recommended_ci_value
    last_insulin_carb = None
    last_insulin_correction = None
    last_total_insulin = None
    recommended_ci_value = ""        


def clear_all():
    if messagebox.askyesno("確認清除", "確定要清除碳水計算項目嗎？此操作無法還原。"):
        search_listbox.delete(0, tk.END)
        calc_listbox.delete(0, tk.END)
        entry_amount.delete(0, tk.END)
        lbl_total.config(text="總碳水量：0g")
        calc_results.clear()
        
def clear_carb_calc_tab():
    if messagebox.askyesno("確認清除", "確定要清除所有碳水計算項目嗎？此操作無法還原。"):
        calc_listbox.delete(0, tk.END)
        entry_amount.delete(0, tk.END)
        calc_results.clear()
        lbl_total.config(text="總碳水量：0g")
    

# === 初始化 ===
init_excel()
current_matches = []
calc_results = []

# === 主視窗 ===
root = tk.Tk()
root.title("食物碳水與胰島素記錄系統")
root.geometry("600x950")  # 調整整體大小

# === 建立分頁框架 ===
tabs = ttk.Notebook(root)
tab_food = ttk.Frame(tabs)
tab_calc = ttk.Frame(tabs)
tab_insulin = ttk.Frame(tabs)
#tab_recommend = ttk.Frame(tabs)

tabs.add(tab_food, text="🍱 食物管理")
tabs.add(tab_calc, text="📊 碳水計算")
tabs.add(tab_insulin, text="💉 血糖與胰島素劑量紀錄")
#tabs.add(tab_recommend, text="🔁 劑量與建議")
tabs.pack(expand=1, fill="both")

# === 食物管理分頁（美化後） ===
frame_food_top = tk.LabelFrame(tab_food, text="🔸 新增或更新食物資料", font=("Arial", 13, "bold"), padx=10, pady=10)
frame_food_top.pack(fill="x", padx=10, pady=(10, 5))

tk.Label(frame_food_top, text="食物名稱", font=("Arial", 11)).grid(row=0, column=0, sticky="e", padx=5, pady=3)
entry_name = tk.Entry(frame_food_top, font=("Arial", 11))
entry_name.grid(row=0, column=1, padx=5, pady=3)

tk.Label(frame_food_top, text="單位", font=("Arial", 11)).grid(row=1, column=0, sticky="e", padx=5, pady=3)
combo_unit = ttk.Combobox(frame_food_top, values=["克(g)", "毫升(ml)"], state="readonly", font=("Arial", 11))
combo_unit.grid(row=1, column=1, padx=5, pady=3)

tk.Label(frame_food_top, text="碳水(每單位)", font=("Arial", 11)).grid(row=2, column=0, sticky="e", padx=5, pady=3)
entry_carb = tk.Entry(frame_food_top, font=("Arial", 11))
entry_carb.grid(row=2, column=1, padx=5, pady=3)

tk.Label(frame_food_top, text="備註", font=("Arial", 11)).grid(row=3, column=0, sticky="e", padx=5, pady=3)
entry_note = tk.Entry(frame_food_top, font=("Arial", 11))
entry_note.grid(row=3, column=1, padx=5, pady=3)

tk.Button(frame_food_top, text="✅ 新增 / 覆蓋食物", bg="#f0ad4e", font=("Arial", 11), command=check_and_save_food).grid(row=4, column=0, columnspan=2, pady=10)

frame_food_query = tk.LabelFrame(tab_food, text="🔍 查詢與刪除", font=("Arial", 13, "bold"), padx=10, pady=10)
frame_food_query.pack(fill="x", padx=10, pady=(5, 10))

entry_search = tk.Entry(frame_food_query, font=("Arial", 11))
entry_search.pack(side="left", padx=5, expand=True, fill="x")

tk.Button(frame_food_query, text="查詢", font=("Arial", 11), command=search_food).pack(side="left", padx=5)

search_listbox = tk.Listbox(tab_food, width=60, height=5, font=("Arial", 11))
search_listbox.pack(padx=10, pady=5)

tk.Button(tab_food, text="🗑️ 刪除食物資料", bg="#d9534f", fg="white", font=("Arial", 11), command=delete_food).pack(pady=5)


# === 📊 碳水計算分頁（美化版） ===

# 1️⃣ 區塊標題
tk.Label(tab_calc, text="📋 食物攝取計算區", font=("Arial", 13, "bold")).pack(pady=(10, 0))

# 2️⃣ 查詢食物區塊
frame_search = tk.LabelFrame(tab_calc, text="🔍 查詢食物", font=("Arial", 11, "bold"), padx=10, pady=10)
frame_search.pack(fill="x", padx=10, pady=10)

entry_search_calc = tk.Entry(frame_search, font=("Arial", 11))
entry_search_calc.pack(side="left", expand=True, fill="x", padx=(0, 10))

btn_search_calc = tk.Button(frame_search, text="查詢", command=search_food_calc, font=("Arial", 11))
btn_search_calc.pack(side="right")

# 顯示查詢結果 Listbox
search_listbox_calc = tk.Listbox(tab_calc, width=60, height=5, font=("Arial", 11))
search_listbox_calc.pack(padx=10, pady=(0, 5))

# 刪除食物按鈕
btn_delete_calc_food = tk.Button(tab_calc, text="🗑️ 刪除食物資料", bg="#d9534f", fg="white", font=("Arial", 11), command=delete_food)
btn_delete_calc_food.pack(pady=(0, 10))

# 3️⃣ 加入攝取量區塊
frame_input = tk.LabelFrame(tab_calc, text="📥 加入攝取與計算", font=("Arial", 11, "bold"), padx=10, pady=10)
frame_input.pack(fill="x", padx=10, pady=(0, 10))

tk.Label(frame_input, text="攝取量 (g/ml)", font=("Arial", 11)).grid(row=0, column=0, sticky="e", padx=5, pady=5)
entry_amount = tk.Entry(frame_input, font=("Arial", 11))
entry_amount.grid(row=0, column=1, sticky="w", padx=5, pady=5)

btn_add = tk.Button(frame_input, text="✅ 加入計算", bg="#5cb85c", fg="white", font=("Arial", 11), command=add_to_calc_calc_tab)
btn_add.grid(row=1, column=0, columnspan=2, pady=(5, 0))

# 顯示已加入食物項目
calc_listbox = tk.Listbox(tab_calc, width=60, height=8, font=("Arial", 11))
calc_listbox.pack(padx=10, pady=(0, 5))

# 刪除選擇項目
btn_delete_calc = tk.Button(tab_calc, text="❌ 刪除選擇項目", bg="#dc3545", fg="white", font=("Arial", 11), command=delete_selected_calc_item)
btn_delete_calc.pack(pady=(0, 5))

# 顯示總碳水量
lbl_total = tk.Label(tab_calc, text="總碳水量：0g", font=("Arial", 11, "bold"), fg="#333")
lbl_total.pack(pady=(0, 10))

btn_clear_carb = tk.Button(
    tab_calc,
    text="🧹 清除碳水項目",
    bg="#dc3545", fg="white",
    font=("Arial", 11), width=20,
    command=clear_carb_calc_tab
)
btn_clear_carb.pack(pady=10)

# === 💉 血糖與胰島素紀錄分頁（美化後） ===

# 分頁標題
tk.Label(tab_insulin, text="📄 血糖與基本設定輸入", font=("Arial", 13, "bold")).pack(pady=(10, 5))

# 設定區塊使用 LabelFrame 包裹
frame_insulin_input = tk.LabelFrame(tab_insulin, text="🔧 基本輸入設定", font=("Arial", 11, "bold"), padx=10, pady=10)
frame_insulin_input.pack(fill="x", padx=15, pady=(5, 10))

# 1️⃣ 日期
tk.Label(frame_insulin_input, text="📅 日期", font=("Arial", 11)).grid(row=0, column=0, sticky="e", padx=8, pady=5)
entry_date = DateEntry(frame_insulin_input, date_pattern="yyyy-mm-dd", font=("Arial", 10))
entry_date.grid(row=0, column=1, padx=8, pady=5)

# 2️⃣ 餐別
tk.Label(frame_insulin_input, text="🍽️ 餐別", font=("Arial", 11)).grid(row=1, column=0, sticky="e", padx=8, pady=5)
combo_meal = ttk.Combobox(frame_insulin_input, values=["早餐", "午餐", "晚餐", "宵夜"], state="readonly", font=("Arial", 10))
combo_meal.grid(row=1, column=1, padx=8, pady=5)

# 3️⃣ 目前血糖值
tk.Label(frame_insulin_input, text="🩸 目前血糖值", font=("Arial", 11)).grid(row=2, column=0, sticky="e", padx=8, pady=5)
entry_current_glucose = tk.Entry(frame_insulin_input, font=("Arial", 11))
entry_current_glucose.grid(row=2, column=1, padx=8, pady=5)

# 4️⃣ 期望血糖值
tk.Label(frame_insulin_input, text="🎯 期望血糖值", font=("Arial", 11)).grid(row=3, column=0, sticky="e", padx=8, pady=5)
entry_target_glucose = tk.Entry(frame_insulin_input, font=("Arial", 11))
entry_target_glucose.grid(row=3, column=1, padx=8, pady=5)

# 5️⃣ 餐後血糖值
tk.Label(frame_insulin_input, text="📈 餐後血糖值", font=("Arial", 11)).grid(row=4, column=0, sticky="e", padx=8, pady=5)
entry_post_glucose = tk.Entry(frame_insulin_input, font=("Arial", 11))
entry_post_glucose.grid(row=4, column=1, padx=8, pady=5)

# 📥 儲存資料按鈕區塊
frame_buttons = tk.Frame(tab_insulin)
frame_buttons.pack(pady=10)

btn_save_record = tk.Button(
    frame_buttons, text="💾 儲存記錄", bg="#0275d8", fg="white", font=("Arial", 11), width=20,
    command=save_records
)
btn_save_record.pack(pady=5)

btn_save_post = tk.Button(
    frame_buttons, text="📥 儲存餐後血糖值", bg="#17a2b8", fg="white", font=("Arial", 11), width=20,
    command=save_post_glucose_only
)
btn_save_post.pack(pady=5)

# === 🔁 建議與計算區塊（合併後） ===
tk.Label(tab_insulin, text="🧮 劑量與建議計算區", font=("Arial", 13, "bold")).pack(pady=(10, 5))

# 📌 基本輸入參數區
frame_inputs = tk.LabelFrame(tab_insulin, text="📌 基本輸入參數", font=("Arial", 11, "bold"), padx=10, pady=10)
frame_inputs.pack(fill="x", padx=15, pady=(5, 10))

labels3 = ["C/I 值", "ISF 值", "1C 升高血糖"]
entries3 = []

for i, text in enumerate(labels3):
    tk.Label(frame_inputs, text=text, font=("Arial", 11)).grid(row=i, column=0, sticky="e", padx=8, pady=5)
    entry = tk.Entry(frame_inputs, font=("Arial", 11))
    entry.grid(row=i, column=1, padx=8, pady=5)
    entries3.append(entry)

entry_ci, entry_isf, entry_c_raise = entries3
entry_isf.insert(0, "50")

# 插入查詢建議 C/I 按鈕
btn_load_ci = tk.Button(
    frame_inputs, text="🔍 載入建議 C/I", font=("Arial", 10),
    command=load_last_recommended_ci
)
btn_load_ci.grid(row=0, column=2, padx=5)


# 🧮 胰島素劑量按鈕與結果
frame_calc_result = tk.Frame(tab_insulin)
frame_calc_result.pack(pady=(5, 10))

btn_calc_insulin = tk.Button(
    frame_calc_result, text="🧮 計算胰島素劑量", bg="#6f42c1", fg="white", font=("Arial", 11),
    width=25, command=calculate_insulin_dose
)
btn_calc_insulin.pack(pady=5)

label_insulin_result = tk.Label(frame_calc_result, text="", font=("Arial", 11, "bold"), fg="#333")
label_insulin_result.pack()

# 📉 建議 C/I 值區塊
frame_ci = tk.LabelFrame(tab_insulin, text="📉 回推建議 C/I 值", font=("Arial", 11, "bold"), padx=10, pady=10)
frame_ci.pack(fill="x", padx=15, pady=(5, 10))

btn_ci_suggest = tk.Button(
    frame_ci, text="📉 計算建議 C/I", bg="#20c997", fg="white", font=("Arial", 11),
    width=25, command=calculate_recommended_ci
)
btn_ci_suggest.pack(pady=5)

label_ci_recommend = tk.Label(frame_ci, text="", font=("Arial", 11, "bold"))
label_ci_recommend.pack(pady=(2, 0))

label_ci_load_info = tk.Label(frame_ci, text="", font=("Arial", 10), fg="green")
label_ci_load_info.pack()

btn_clear_form = tk.Button(
    frame_buttons, text="🧹 清除輸入欄位", bg="#6c757d", fg="white", font=("Arial", 11), width=20,
    command=clear_insulin_tab
)
btn_clear_form.pack(pady=5)

# === 綁定分頁切換事件 ===
tabs.bind("<<NotebookTabChanged>>", on_tab_recommend_selected)

# === 進入主事件迴圈 ===
root.mainloop()