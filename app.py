# Streamlit Web App for Food Carb & Insulin Record System (Excel-based)
# 🔧 初步版本，保留三分頁架構與原 Excel 資料互動邏輯

import streamlit as st
from openpyxl import load_workbook, Workbook
import pandas as pd
import os
from datetime import datetime
from fuzzywuzzy import fuzz

FOOD_FILE = "foodssugar.xlsx"
RECORD_FILE = "Ruby_records.xlsx"

st.set_page_config(page_title="食物碳水與胰島素系統", layout="wide")

# === 初始化 Excel ===
def init_excel():
    if not os.path.exists(FOOD_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "食物資料"
        ws.append(["食物名稱", "單位", "碳水化合物", "備註"])
        wb.save(FOOD_FILE)

    if not os.path.exists(RECORD_FILE):
        wb = Workbook()
        wb.create_sheet("食物記錄")
        wb.create_sheet("血糖與胰島素紀錄表")
        wb.remove(wb["Sheet"])
        wb.save(RECORD_FILE)

# === 食物查詢 ===
def find_similar_foods(keyword, threshold=80):
    if not os.path.exists(FOOD_FILE):
        return []
    wb = load_workbook(FOOD_FILE)
    ws = wb.active
    matches = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if fuzz.partial_ratio(keyword, str(row[0])) >= threshold:
            matches.append(row)
    return matches

# === 儲存紀錄 ===
def save_insulin_record(date, meal, total_carb, current_glucose, target_glucose, ci, isf, insulin_carb, insulin_correction, total_insulin):
    if os.path.exists(RECORD_FILE):
        wb = load_workbook(RECORD_FILE)
    else:
        wb = Workbook()
        wb.create_sheet("血糖與胰島素紀錄表")
        wb.remove(wb["Sheet"])

    if "血糖與胰島素紀錄表" not in wb.sheetnames:
        ws = wb.create_sheet("血糖與胰島素紀錄表")
        ws.append([
            "日期", "餐別", "總碳水量", "目前血糖值", "期望血糖值",
            "C/I值", "ISF值", "碳水劑量", "矯正劑量", "總胰島素劑量"
        ])
    else:
        ws = wb["血糖與胰島素紀錄表"]

    ws.append([
        date, meal, total_carb, current_glucose, target_glucose,
        ci, isf, insulin_carb, insulin_correction, total_insulin
    ])
    wb.save(RECORD_FILE)

# === 初始化 ===
init_excel()

# === Session State 初始化 ===
if "calc_results" not in st.session_state:
    st.session_state.calc_results = []

# === 分頁設定 ===
tabs = st.tabs(["🍱 食物管理", "📊 碳水計算", "💉 胰島素紀錄"])

# === 1. 食物管理 ===
with tabs[0]:
    st.header("🍱 食物管理")
    st.subheader("新增或查詢食物")

    with st.form("add_food_form"):
        name = st.text_input("食物名稱")
        unit = st.selectbox("單位", ["克(g)", "毫升(ml)"])
        carb = st.text_input("每單位碳水化合物含量 (g)")
        note = st.text_input("備註")
        submitted = st.form_submit_button("✅ 新增 / 覆蓋")

        if submitted:
            if not name or not unit or not carb:
                st.warning("請填寫完整資訊")
            else:
                try:
                    carb_val = float(carb.replace(",", "."))
                    matches = find_similar_foods(name)
                    wb = load_workbook(FOOD_FILE)
                    ws = wb.active
                    updated = False
                    if matches:
                        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
                            if fuzz.partial_ratio(name, row[0].value) >= 80:
                                ws.cell(i, 1, name)
                                ws.cell(i, 2, unit)
                                ws.cell(i, 3, carb_val)
                                ws.cell(i, 4, note)
                                updated = True
                                break
                    if not updated:
                        ws.append([name, unit, carb_val, note])
                    wb.save(FOOD_FILE)
                    st.success("✅ 已儲存：{}".format("覆蓋" if updated else "新增"))
                except:
                    st.error("碳水值請輸入數字")

    st.divider()
    st.subheader("🔍 查詢 / 刪除食物")
    keyword = st.text_input("查詢關鍵字")
    if keyword:
        results = find_similar_foods(keyword)
        if results:
            df = pd.DataFrame(results, columns=["食物名稱", "單位", "碳水(g)", "備註"])
            st.dataframe(df, use_container_width=True)
            selected = st.selectbox("選擇要刪除的項目", [r[0] for r in results])
            if st.button("🗑️ 刪除選擇項目"):
                wb = load_workbook(FOOD_FILE)
                ws = wb.active
                for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
                    if row[0].value == selected:
                        ws.delete_rows(i)
                        wb.save(FOOD_FILE)
                        st.success(f"已刪除：{selected}")
                        break
        else:
            st.info("查無資料，請確認輸入內容或先新增食物")

# === 2. 碳水計算 ===
with tabs[1]:
    st.header("📊 碳水化合物攝取計算")
    st.subheader("查詢食物並輸入攝取量")

    col1, col2 = st.columns([2, 1])
    with col1:
        keyword_calc = st.text_input("輸入食物名稱以查詢")
    with col2:
        search = st.button("🔍 查詢")

    calc_matches = []
    if search and keyword_calc:
        calc_matches = find_similar_foods(keyword_calc)
        if not calc_matches:
            st.warning("查無資料，請確認輸入或先新增食物")

    if calc_matches:
        selected = st.selectbox("選擇食物項目", [f"{r[0]}｜每{r[1]} 含 {r[2]}g" for r in calc_matches])
        amount = st.number_input("攝取量 (g/ml)", min_value=0.0, step=1.0)
        if st.button("✅ 加入計算"):
            idx = [f"{r[0]}｜每{r[1]} 含 {r[2]}g" for r in calc_matches].index(selected)
            row = calc_matches[idx]
            carb = round(float(row[2]) * amount, 2)
            st.session_state.calc_results.append({"name": row[0], "amount": amount, "unit": row[1], "carb": carb})
            st.success(f"已加入：{row[0]}｜{amount}{row[1]}｜碳水: {carb}g")

    st.divider()
    st.subheader("📋 已加入項目")
    if st.session_state.calc_results:
        df_calc = pd.DataFrame(st.session_state.calc_results)
        df_calc.columns = ["食物", "攝取量", "單位", "碳水(g)"]
        st.dataframe(df_calc, use_container_width=True)
        total = sum([r["carb"] for r in st.session_state.calc_results])
        st.metric("總碳水量 (g)", f"{round(total, 2)}")
        if st.button("❌ 清除所有項目"):
            st.session_state.calc_results.clear()
            st.success("已清除")
    else:
        st.info("尚未加入任何食物項目")

# === 3. 胰島素紀錄 ===
with tabs[2]:
    st.header("💉 胰島素劑量紀錄與建議")
    st.subheader("輸入血糖資訊與參數")

    col1, col2 = st.columns(2)
    with col1:
        date = st.date_input("📅 日期", value=datetime.today())
        meal = st.selectbox("🍽️ 餐別", ["早餐", "午餐", "晚餐", "宵夜"])
        current_glucose = st.number_input("🩸 目前血糖值", min_value=0, step=1)
        target_glucose = st.number_input("🎯 期望血糖值", min_value=0, value=100)
    with col2:
        ci = st.number_input("C/I 值 (每U能代謝幾克碳水)", min_value=0.1, step=0.1)
        isf = st.number_input("ISF 值 (每U能降低幾 mg/dL)", min_value=0.1, step=0.1)

    st.divider()
    st.subheader("🧮 劑量計算結果")

    if st.button("🧮 計算胰島素劑量"):
        total_carb = round(sum([r["carb"] for r in st.session_state.calc_results]), 2)
        insulin_carb = round(total_carb / ci, 1)
        insulin_correction = round((current_glucose - target_glucose) / isf, 1)
        total_insulin = round(insulin_carb + insulin_correction, 1)

        st.session_state.last_insulin_result = {
            "insulin_carb": insulin_carb,
            "insulin_correction": insulin_correction,
            "total_insulin": total_insulin,
            "total_carb": total_carb
        }

        st.success(f"碳水劑量：{insulin_carb}U\n矯正劑量：{insulin_correction}U\n總劑量：{total_insulin}U")

    if "last_insulin_result" in st.session_state:
        if st.button("💾 儲存紀錄"):
            r = st.session_state.last_insulin_result
            save_insulin_record(
                str(date), meal, r["total_carb"], current_glucose,
                target_glucose, ci, isf,
                r["insulin_carb"], r["insulin_correction"], r["total_insulin"]
            )
            st.success("✅ 已儲存紀錄至 Excel")
